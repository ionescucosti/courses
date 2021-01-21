# 816
# Excel openpyxl - manipularea in memorie
# PF - 12.01.2017

# instalare modul:      pip install openpyxl
#                       pip install xlwt - pentru fisiere .xls - exemplele sunt pt .xlsx
# instalare modul imagini:      pip install pillow
# documentatie : https://openpyxl.readthedocs.io/en/default/     next urmatoarele pagini

from openpyxl import Workbook,load_workbook
wb = Workbook()		# Creeaza obiectul care sustine fisierul excel

ws = wb.active		# Lucreaza in worksheet-ul curent (ws) - il creeaza daca nu exista
                    # O foaie creata nu contine campuri. Ele sunt create pe masura ce populam fisierul

for i in range(1,101):
    for j in range(1,101):
        ws.cell(row=i, column=j)    # Asigneaza campuri unei foi in memorie 100x100

ws1 = wb.create_sheet()     # Daca nu specificam numele implicit sheet, sheet1, etc
ws2 = wb.create_sheet('Paul', 1)    # Atribuie numele si poozitia in workbook
ws3 = wb.create_sheet('Jeni')       # Atribuie numele, ultimul sheet
ws4 = wb.create_sheet('Fane', 0)    # Atribuie numele, primul sheet

ws1.title = 'Ioana'     # Schimba numele sheet-ului

ws.sheet_properties.tabColor = "1072BA"     # background color RRGGBB color code

ws5 = wb['Ioana']       # Atribuie un alias pentru foaia cu numele dat

print(wb.sheetnames)    # Printeaza o lista cu foile existente

for sheet in wb:
    print(sheet.title)  # Printeaza foile existente

ws['A1'] = 55		    # Populare camp direct. Daca exista il rescrie

ws.append([7, 8, 9])	# Adaugare pe primele 3 coloane urmatorul rand liber

source = wb.active
target = wb.copy_worksheet(source)  # Creeaza o copie a foii curente

var = ws['C1']                              # Variabila corespunzatoare campului
var1 = ws.cell(row=4, column=2, value=10)   # Variabila corespunzatoare campului cu asignare

var.value = 101         # Modificare valoare camp
var1.value = 'pisica'   # Modificare valoare camp

print(var1.value)        # printeaza valoarea celulei

print(ws['B4'].value)   # printeaza valoarea celulei. Poate fi folosita in expresii

cell_range = ws['A1':'B4']      # range de campuri
colC = ws['a']                  # o coloana
col_range = ws['C:D']           # range de coloane
row10 = ws[10]                  # un rand
row_range = ws[5:10]            # range de randuri

for row in ws.iter_rows(min_row=1, max_col=3, max_row=2):
    for cell in row:
        print(cell)             # for in fisier

tuple(ws.rows)          # returneaza tupluri cu celulele fiecarui rand

tuple(ws.columns)       # returneaza tupluri cu celulele fiecarei coloane

ws.append([7, 8, 9])

import datetime
ws['F5'] = datetime.datetime.now()		# Populeaza data si ora, transforma automat in formatul excel

wb.save("sample.xlsx")   # Salvarea propriu-zisa a fisierului. Daca exista il rescrie
                           # Poate fi salvat cu .xlsx sau .xlsm

wb1 = load_workbook('sample16.xlsx')    # putem deschide un fisier existent
wb1.template = True                     # daca setam false salvarea se face ca document normal xlsx
wb1.save('sample_t.xltx')               # putem crea un template

"""There are several flags that can be used in      load_workbook.

guess_types will enable or disable (default) type inference when reading cells.
data_only controls whether cells with formulae have either the formula (default)
or the value stored the last time Excel read the sheet.
keep_vba controls whether any Visual Basic elements are preserved or not (default).
If they are preserved they are still not editable.
"""

print(wb1.sheetnames)
print(wb1.get_sheet_names())

for i in range(1,101):
    for j in range(1,101):
        if ws.cell(row=i, column=j).value:
            print(ws.cell(row=i, column=j).value)   # returneaza valorile din campuri

ws['A1'].number_format      # tip de date numeric

ws["A3"] = "=SUM(1, 7)"     # lucru cu formule

