# 816
# Excel openpyxl - manipularea in memorie
# PF - 12.01.2017

# instalare modul:      pip3 install openpyxl
#                       pip3 install xlwt - pentru fisiere .xls - exemplele sunt pt .xlsx
# instalare modul imagini:      pip3 install pillow
# documentatie : https://openpyxl.readthedocs.io/en/default/     next urmatoarele pagini

import csv

from openpyxl import Workbook

wb = Workbook()  # Creeaza obiectul care sustine fisierul excel

ws = wb.active  # Lucreaza in worksheet-ul curent (ws) - il creeaza daca nu exista
# O foaie creata nu contine campuri. Ele sunt create pe masura ce populam fisierul

# for i in range(1,256):
#   for j in range(1,256):
#      ws.cell(row=i, column=j)    # Asigneaza campuri unei foi in memorie 100x100


with open('c:/wt/medici.csv', newline='', encoding='utf8') as csvfile:
    medici_reader = csv.reader(csvfile, delimiter=',', quotechar=None)
    indx = 0
    for row in medici_reader:
        ws.append(row)
        indx += 1

with open('c:/wt/medici.csv', newline='', encoding='utf8') as csvfile:
    medici_reader = csv.reader(csvfile, delimiter=',', quotechar=None)
    indx = 0
    for row in medici_reader:
        ws.append(row + [indx])
        indx += 1

with open('c:/wt/medici.csv', newline='', encoding='utf8') as csvfile:
    medici_reader = csv.reader(csvfile, delimiter=',', quotechar=None)
    indx = 0
    for row in medici_reader:
        if row[3] == 'o.r.l.':
            ws.append(row + [indx])
            indx += 1


wb1 = Workbook()

ws1 = wb1.active

for i in ws['a']:


wb.save("sample19.xlsx")  # Salvarea propriu-zisa a fisierului. Daca exista il rescrie
# Poate fi salvat cu .xlsx sau .xlsm
